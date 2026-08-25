#!/usr/bin/env python3
"""Convert a Crunchbase Pro company export (.xlsx) into the compact JSON
universe bundled at data/universe-crunchbase.json.

Usage: python3 scripts/convert-crunchbase.py <export.xlsx>

Keeps only Active companies and the fields the match engine reads:
firmographics (employee band, estimated revenue band, funding stage),
dated timing events (last funding, last leadership hire, last layoff
mention), growth category, and people-graph-lite flags derived from the
Current/Past Employees rosters (which GTM functions are present/departed).
"""
import sys, json, re
from datetime import datetime
import openpyxl

SL = re.compile(r'vp of sales|chief revenue|head of sales|\bcro\b|sales director', re.I)
MK = re.compile(r'vp of marketing|\bcmo\b|chief marketing|head of marketing|marketing director', re.I)
OPS = re.compile(r'revenue operations|revops|sales operations|marketing operations', re.I)
CS = re.compile(r'customer success|chief customer', re.I)
EN = re.compile(r'enablement', re.I)
PT = re.compile(r'partnership|alliances', re.I)
AI = re.compile(r'machine learning|ml engineer|ai engineer|data scientist', re.I)
ROLES = {'sl': SL, 'mk': MK, 'ops': OPS, 'cs': CS, 'en': EN, 'pt': PT, 'ai': AI}

STAGE_FROM_EQUITY = {'pre-seed': 'pre_seed', 'angel': 'pre_seed', 'seed': 'seed',
                     'series a': 'series_a', 'series b': 'series_b'}
STAGE_FROM_STATUS = {'seed': 'seed', 'early stage venture': 'series_a',
                     'late stage venture': 'series_c_plus', 'private equity': 'growth',
                     'm&a': 'growth', 'ipo': 'growth'}
EMP_BANDS = {'1-10': ('1_10', 5), '11-50': ('11_50', 30), '51-100': ('51_200', 75),
             '101-250': ('51_200|201_500', 175), '251-500': ('201_500', 375),
             '501-1000': ('501_1000', 750), '1001-5000': ('1001_plus', 3000),
             '5001-10000': ('1001_plus', 7500), '10001+': ('1001_plus', 15000)}
REV_MAP = {'less than $1m': 'under_1m', '$1m to $10m': '1m_5m|5m_20m',
           '$10m to $50m': '5m_20m|20m_50m', '$50m to $100m': '50m_plus',
           '$100m to $500m': '50m_plus', '$500m to $1b': '50m_plus',
           '$1b to $10b': '50m_plus', '$10b+': '50m_plus'}


def main(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    idx = {h: i for i, h in enumerate(next(rows))}

    def s(row, col):
        v = row[idx[col]] if col in idx else None
        return str(v).strip() if v is not None else ""

    def g(row, col):
        return row[idx[col]] if col in idx else None

    def iso(v):
        return v.strftime('%Y-%m-%d') if isinstance(v, datetime) else None

    out = []
    for row in rows:
        if not s(row, 'Organization Name') or s(row, 'Operating Status').lower() != 'active':
            continue
        m = re.match(r'https?://(?:www\.)?([^/]+)', s(row, 'Website'))
        roster = s(row, 'Current Employees') + ' | ' + s(row, 'Contact Job Titles')
        past = s(row, 'Past Employees')
        equity = s(row, 'Last Equity Funding Type').lower()
        stage = next((v for k, v in STAGE_FROM_EQUITY.items() if k in equity), None)
        if not stage and 'series' in equity:
            stage = 'series_c_plus'
        if not stage:
            stage = STAGE_FROM_STATUS.get(s(row, 'Funding Status').lower())
        emp = EMP_BANDS.get(s(row, 'Number of Employees'))
        founded = g(row, 'Founded Date')
        rec = {
            'n': s(row, 'Organization Name'),
            'd': m.group(1).lower() if m else None,
            'o': s(row, 'Description')[:180],
            'i': (s(row, 'Industry Groups') + '; ' + s(row, 'Industries'))[:160],
            'eb': emp[0] if emp else None,
            'ts': emp[1] if emp else None,
            'rv': REV_MAP.get(s(row, 'Estimated Revenue Range').lower()),
            'st': stage,
            'fy': founded.year if isinstance(founded, datetime) else None,
            'fd': iso(g(row, 'Last Funding Date')),
            'ft': s(row, 'Last Funding Type') or None,
            'fa': int(g(row, 'Last Funding Amount')) if g(row, 'Last Funding Amount') else None,
            'h': s(row, 'Actively Hiring') == 'Yes',
            'gc': s(row, 'Growth Category') or None,
            'lh': iso(g(row, 'Last Leadership Hiring Date')),
            'll': iso(g(row, 'Last Layoff Mention Date')),
            'li': s(row, 'LinkedIn') or None,
            'rp': ','.join(k for k, rx in ROLES.items() if rx.search(roster)) or None,
            'rd': ','.join(k for k, rx in ROLES.items() if rx.search(past)) or None,
        }
        out.append({k: v for k, v in rec.items() if v is not None and v != ''})

    with open('data/universe-crunchbase.json', 'w') as f:
        json.dump(out, f, separators=(',', ':'))
    print(f"wrote {len(out)} companies to data/universe-crunchbase.json")


if __name__ == '__main__':
    main(sys.argv[1])
